import os
from pathlib import Path
from openpyxl import Workbook
import parser

dict_of_tags = {"НАЗВАНИЕ ФОНДА":"",
                "УПРАВЛЯЮЩАЯ КОМПАНИЯ": "field_funds_comp_name js_swtch_cntrl_visible",
                "ДАТА РАСЧЕТА": "field_nav_date middle js_swtch_cntrl_visible",
                "СЧА, МЛН RUB":"field_nav js_swtch_cntrl_visible",
                "ПАЙЩИКИ":"field_shareholders_count js_swtch_cntrl_visible",
                "МИНИМАЛЬНЫЙ ВЗНОС, RUB":"field_min_invest js_swtch_cntrl_visible",
                "НАЧАЛО РАБОТЫ":"field_date_of_end_placing js_swtch_cntrl_visible",
                "ТИП":"field_funds_types js_swtch_cntrl_visible",
                "ОБЪЕКТ ИНВЕСТИРОВАНИЯ":"field_funds_object_name js_swtch_cntrl_visible",
                "СТАТУС":"field_funds_statuses_name js_swtch_cntrl_visible",
                "КАТЕГОРИЯ":"field_funds_categories_title js_swtch_cntrl_visible",
                "НАПРАВЛЕНИЕ ИНВЕСТИРОВАНИЯ":"field_funds_investing_directions_name js_swtch_cntrl_visible",
                "УПРАВЛЯЮЩИЙ":"field_funds_contacts js_swtch_cntrl_visible",
                "ВОЗНАГРАЖДЕНИЕ УК":"field_fee_uk js_swtch_cntrl_visible",
                "МАКС. РАСХОДЫ НА УПРАВЛЕНИЕ":"field_fee_total js_swtch_cntrl_visible",
                "ДОХОДНОСТЬ ЗА 1 МЕСЯЦ":"field_delta_pay_1m js_swtch_cntrl_visible",
                "ДОХОДНОСТЬ ЗА 3 МЕСЯЦ":"field_delta_pay_3m js_swtch_cntrl_visible",
                "ДОХОДНОСТЬ С НАЧАЛА ГОДА":"field_delta_pay_ys js_swtch_cntrl_visible",
                "ДОХОДНОСТЬ ЗА 1 ГОД":"field_delta_pay_1y js_swtch_cntrl_visible",
                "ДОХОДНОСТЬ ЗА 3 ГОДА":"field_delta_pay_3y js_swtch_cntrl_visible",
                "ДОХОДНОСТЬ ЗА 5 ЛЕТ":"field_delta_pay_5y js_swtch_cntrl_visible",
                "ПРИВЛЕЧЕННЫЕ СРЕДСТВА ЗА 1 МЕСЯЦ, МЛН RUB":"field_volume_opc_1m js_swtch_cntrl_visible",
                "ПРИВЛЕЧЕННЫЕ СРЕДСТВА ЗА 3 МЕСЯЦА, МЛН RUB":"field_volume_opc_3m js_swtch_cntrl_visible",
                "ПРИВЛЕЧЕННЫЕ СРЕДСТВА С НАЧАЛА ГОДА, МЛН RUB":"field_volume_opc_ys js_swtch_cntrl_visible",
                "ПРИВЛЕЧЕННЫЕ СРЕДСТВА ЗА 1 ГОД, МЛН RUB":"field_volume_opc_1y js_swtch_cntrl_visible",
                "ПРИВЛЕЧЕННЫЕ СРЕДСТВА ЗА 3 ГОДА, МЛН RUB":"field_volume_opc_3y js_swtch_cntrl_visible",
                "ПРИВЛЕЧЕННЫЕ СРЕДСТВА ЗА 5 ЛЕТ, МЛН RUB":"field_volume_opc_5y js_swtch_cntrl_visible",
                "ДАТА КОЭФФИЦИЕНТОВ":"field_coeff_date js_swtch_cntrl_visible",
                "ШАРПА":"field_coeff_sharp js_swtch_cntrl_visible",
                "СОРТИНО":"field_coeff_srtn js_swtch_cntrl_visible",
                "ВОЛАТИЛЬНОСТЬ":"field_coeff_vol js_swtch_cntrl_visible",
                "АЛЬФА":"field_coeff_alpha js_swtch_cntrl_visible",
                "БЕТА":"field_coeff_beta js_swtch_cntrl_visible",
                "R2":"field_coeff_r2 js_swtch_cntrl_visible",
                "VAR":"field_coeff_var js_swtch_cntrl_visible"
                }


def main_excel(dir_path):
    for page in range(1, 13):
        wb = Workbook()
        if not os.path.exists(dir_path):
            os.mkdir(dir_path)
        url = f"https://investfunds.ru/funds/?showID=99&cstm=0-3y283xa.1-2&cmp=0-42334.1-zik0zo.2-74.3-1ekg.5-9zlds.41-k1zwg.49-6bk.62-1kw.112-6bk.157-2t4w&limit=50&page={page}&sortId=99"
        main_driver = parser.driver_init(url)
        ws = wb.active
        add_header(ws)
        add_first_column(main_driver, ws)
        main_loop(main_driver, ws)
        main_driver.close()
        wb.save(Path(dir_path) / f"invest{page}.xlsx")


def add_header(ws):
    dict = list(dict_of_tags)
    counter = 0
    for col in ws.iter_cols(max_row=1, max_col=len(dict)):
        for cell in col:
            ws.cell(column=cell.column, row=cell.row).value = dict[counter]
            counter += 1


def add_first_column(driver, sheet):
    html = parser.get_html(driver)
    lst = parser.get_cleaned_elements_from_first_column(html)
    counter = 0
    for row in sheet.iter_rows(max_col=1, min_row=2, max_row=len(lst) + 1):
        for cell in row:
            sheet.cell(column=cell.column, row=cell.row).value = lst[counter]
            counter += 1


def add_main_table(driver, sheet, classname_, num_col):
    html = parser.get_html(driver)
    lst = parser.get_cleaned_elements_from_main_table(html, classname=classname_)
    counter = 0
    for row in sheet.iter_rows(min_col=num_col, max_col=num_col, min_row=2, max_row=len(lst) + 1):
        for cell in row:
            try:
                sheet.cell(column=cell.column, row=cell.row).value = lst[counter]
                counter += 1
            except IndexError:
                break


def main_loop(driver, sheet):
    for i, k in enumerate(dict_of_tags, start=2):
        add_main_table(driver, sheet, dict_of_tags[k], i)


if __name__ == "__main__":
    pass


